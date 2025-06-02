# function.py
from PIL import Image
import io
import concurrent.futures
import re
import base64
import requests
import shutil
import os
import json
import threading
# Note: 'sys' and 'datetime' from original full script are not directly used in these functions.
# 'QApplication' from PyQt5 is not used here.

# Added for Excel generation
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None # Flag that openpyxl is not available


def sanitize_filename(filename):
    """
    清除文件名中非法的字符，Windows系统中不允许出现下列字符：\ / : * ? " < > |
    """
    filename = filename.strip().strip('"')
    return re.sub(r'[\\/:*?"<>|]', '', filename)

def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def remove_punctuation_at_end(sentence):
    return re.sub(r'[。？！，、；：“”‘’《》（）【】『』「」\[\]\.,;:"\'?!(){}<>]+$', '', sentence)

def compress_and_encode_image(image_path, quality=85, max_size=(1080, 1080)):
    try:
        with Image.open(image_path) as img:
            img.thumbnail(max_size)
            if img.mode in ('RGBA', 'LA', 'P') and (img.format != 'JPEG' and img.format != 'WEBP'):
                if img.format != 'PNG':
                    img = img.convert('RGB')

            if img.format == 'WEBP' or img.mode == 'RGBA' or img.format == 'GIF':
                output_format = 'PNG'
                mime_type = 'image/png'
            else:
                output_format = img.format if img.format != 'JPEG' else 'JPEG'
                mime_type = f'image/{output_format.lower()}'

            img_bytes = io.BytesIO()
            save_params = {}
            if output_format == 'JPEG':
                save_params['quality'] = quality
                save_params['progressive'] = True
                save_params['optimize'] = True
            elif output_format == 'PNG':
                save_params['optimize'] = True
            elif output_format == 'WEBP':
                 save_params['quality'] = quality

            if output_format == 'JPEG' and img.mode == 'RGBA':
                img_to_save = img.convert('RGB')
                img_to_save.save(img_bytes, format=output_format, **save_params)
            else:
                img.save(img_bytes, format=output_format, **save_params)

            img_bytes.seek(0)
            base64_encoded = base64.b64encode(img_bytes.read()).decode('utf-8')
        return base64_encoded, mime_type, output_format
    except FileNotFoundError:
        raise
    except Exception as e:
        raise RuntimeError(f"Error during image compression/encoding for {image_path}: {e}")


def get_unique_filename(filepath):
    base, ext = os.path.splitext(filepath)
    counter = 1
    new_filepath = filepath
    while os.path.exists(new_filepath):
        new_filepath = f"{base}_{counter}{ext}"
        counter += 1
    return new_filepath

class Counter:
    def __init__(self):
        self.value = 0
        self.lock = threading.Lock()
    def increment(self):
        with self.lock:
            self.value += 1
    def decrement(self):
        with self.lock:
            self.value -= 1
    def get_value(self):
        with self.lock:
            return self.value

# Modified to include renaming_data_list and ensure all attempts are logged for Excel
def process_image(image_path, config, output_text_signal_emit, stop_event, success_counter, failure_counter, num_counter, active_counter, renaming_data_list):
    current_original_filename = os.path.basename(image_path)
    try:
        if stop_event.is_set():
            return # Not processed, so not added to Excel as "failed"

        active_counter.increment()

        api_key = config['Api_key']
        base_url = (
            f"{config['Base_url'].strip().rstrip('/')}/v1/chat/completions"
            if config.get('Base_url', '').strip()
            else "https://yunwu.ai/v1/chat/completions"
        )
        prompt = config['Prompt']
        image_quality_percent = int(config.get('Image_quality_percent', 85))
        quality_value = min(95, max(1, image_quality_percent))

        original_format = os.path.splitext(image_path)[1][1:].upper()
        if not original_format: original_format = "PNG" # Default format

        encoded_image, mime_type, image_format = compress_and_encode_image(image_path, quality=quality_value, max_size=(512, 512))

        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        data = {
            "model": config["Model"],
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{encoded_image}"}}
                    ]
                }
            ],
            "max_tokens": 300
        }

        response = requests.post(base_url, headers=headers, json=data, timeout=60)
        response.raise_for_status() # Will raise HTTPError for bad responses (4xx or 5xx)
        response_data = response.json()

        if 'choices' in response_data and len(response_data['choices']) > 0 and \
           isinstance(response_data['choices'][0].get('message'), dict) and \
           response_data['choices'][0]['message'].get('content'):
            result_text = response_data['choices'][0]['message']['content']
            result_text = sanitize_filename(result_text)

            renaming_data_list.append({'original_name': current_original_filename, 'new_name_suggestion': result_text})

            new_name = f"{result_text}.{original_format.lower()}"
            output_mode = config.get('output_mode', 'finish_subfolder')

            if output_mode == 'custom':
                target_dir = config.get('custom_output_folder', '')
                if not target_dir:
                    output_text_signal_emit(f"错误：自定义输出文件夹未在配置中正确设置。跳过 {current_original_filename}")
                    renaming_data_list.append({'original_name': current_original_filename, 'new_name_suggestion': ''}) # Ensure logged
                    failure_counter.increment(); num_counter.increment(); active_counter.decrement(); return
            elif output_mode == 'in_place':
                target_dir = os.path.dirname(image_path)
            else: # finish_subfolder
                target_dir = os.path.join(os.path.dirname(image_path), 'Finish')

            if not os.path.exists(target_dir):
                try:
                    os.makedirs(target_dir, exist_ok=True)
                except OSError as e:
                    output_text_signal_emit(f"创建文件夹失败 '{target_dir}': {e}. 跳过 {current_original_filename}")
                    # Already added to renaming_data_list if API was successful, if not, added in API error block
                    # If API was success, but folder creation failed, the suggestion was made.
                    # To mark it as truly failed for renaming, we might need to update the entry or handle differently.
                    # For now, if API success, it's logged with suggestion, even if file op fails.
                    # To ensure it's blank if file op fails:
                    # We need to ensure renaming_data_list has an empty suggestion if this part fails.
                    # However, the request is about "生成失败的图片" (generation failed, i.e. AI suggestion).
                    # If AI suggestion is there, but file op fails, the suggestion still exists.
                    # This logic primarily targets AI failure.
                    failure_counter.increment(); num_counter.increment(); active_counter.decrement(); return


            target_path = os.path.join(target_dir, new_name)
            target_path = get_unique_filename(target_path)

            operation_verb = ""
            if output_mode == 'in_place':
                shutil.move(image_path, target_path)
                operation_verb = "在原位置重命名"
            else:
                shutil.copy2(image_path, target_path)
                operation_verb = "复制并重命名到新位置"

            success_counter.increment()
            num_counter.increment()
            output_text_signal_emit(f"第{num_counter.get_value()}张图片处理完成：{current_original_filename} 已{operation_verb}为 {target_dir} (目录: {os.path.basename(target_dir)})")
        else:
            failure_counter.increment()
            num_counter.increment()
            output_text_signal_emit(f"API响应无效或内容为空 ({current_original_filename}): {response_data.get('error', response_data)}")
            renaming_data_list.append({'original_name': current_original_filename, 'new_name_suggestion': ''})

    except FileNotFoundError:
        failure_counter.increment(); num_counter.increment()
        output_text_signal_emit(f"错误：文件未找到 {current_original_filename}")
        renaming_data_list.append({'original_name': current_original_filename, 'new_name_suggestion': ''})
    except requests.exceptions.RequestException as e: # Includes HTTPError from response.raise_for_status()
        failure_counter.increment(); num_counter.increment()
        output_text_signal_emit(f"HTTP请求失败 ({current_original_filename}): {e}")
        renaming_data_list.append({'original_name': current_original_filename, 'new_name_suggestion': ''})
    except RuntimeError as e: # Custom error from compress_and_encode_image
        failure_counter.increment(); num_counter.increment()
        output_text_signal_emit(f"图像处理内部错误 ({current_original_filename}): {e}")
        renaming_data_list.append({'original_name': current_original_filename, 'new_name_suggestion': ''})
    except Exception as e:
        failure_counter.increment(); num_counter.increment()
        output_text_signal_emit(f"处理图片时发生未知错误 ({current_original_filename}): {e}")
        renaming_data_list.append({'original_name': current_original_filename, 'new_name_suggestion': ''})
    finally:
        active_counter.decrement()


def process_images_concurrently(config, output_text_signal_emit, stop_event, active_counter, gui_num_counter):
    source_folder = config["Source_folder"]
    if not source_folder or not os.path.isdir(source_folder):
        output_text_signal_emit(f"错误：源文件夹 '{source_folder}' 无效或未设置。")
        return 0, 0, 0, []

    suffix_name = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.heif', '.heic', '.svg')
    image_paths = [os.path.join(source_folder, filename)
                   for filename in os.listdir(source_folder)
                   if filename.lower().endswith(suffix_name) and os.path.isfile(os.path.join(source_folder, filename))]

    if not image_paths:
        output_text_signal_emit("提示：在源文件夹中没有找到符合条件的图片文件。")
        return 0, 0, 0, []

    success_counter = Counter()
    failure_counter = Counter()
    renaming_data_for_excel = []

    output_text_signal_emit(f"开始处理{len(image_paths)}张图片，线程数: {config.get('Max_workers', 5)}")
    output_text_signal_emit(f"模型：{config['Model']}")
    output_text_signal_emit(f"图片质量设置：{config.get('Image_quality_percent', 85)}%")
    output_mode_display = {
        'finish_subfolder': "保存在 'Finish' 子文件夹",
        'in_place': "在原位置重命名",
        'custom': "保存到自定义文件夹"
    }
    output_text_signal_emit(f"输出模式：{output_mode_display.get(config.get('output_mode', 'finish_subfolder'), '未知')}")
    if config.get('output_mode') == 'custom':
        output_text_signal_emit(f"自定义输出文件夹：{config.get('custom_output_folder')}")

    with concurrent.futures.ThreadPoolExecutor(max_workers=config.get("Max_workers", 5)) as executor:
        futures = {executor.submit(process_image, image_path, config, output_text_signal_emit, stop_event,
                                     success_counter, failure_counter, gui_num_counter, active_counter,
                                     renaming_data_for_excel
                                     ): image_path
                   for image_path in image_paths}

        try:
            for future in concurrent.futures.as_completed(futures):
                if stop_event.is_set():
                    for f_cancel in futures:
                        if not f_cancel.done(): f_cancel.cancel()
                    output_text_signal_emit("停止信号已接收，正在取消剩余任务...")
                    break
                try:
                    future.result()
                except concurrent.futures.CancelledError:
                    output_text_signal_emit("一个任务被取消。") # These won't be in Excel unless process_image added them before cancel
                except Exception:
                    pass # Errors are handled and logged within process_image, and added to renaming_data_for_excel
        except KeyboardInterrupt:
            output_text_signal_emit("检测到键盘中断！正在尝试停止...")
            stop_event.set()
            for f_cancel in futures:
                 if not f_cancel.done(): f_cancel.cancel()
            executor.shutdown(wait=True) # Consider wait=False if UI needs to be more responsive during forced shutdown
            output_text_signal_emit("所有活动线程已尝试停止。")

    return len(image_paths), success_counter.get_value(), failure_counter.get_value(), renaming_data_for_excel


def generate_excel_report(renaming_data, output_folder_path, report_filename="商品标题.xlsx", output_text_signal_emit=None):
    """
    Generates an Excel report with original and new (suggested) filenames.
    """
    if openpyxl is None:
        if output_text_signal_emit:
            output_text_signal_emit("错误：Excel报告生成失败。需要安装 'openpyxl' 库。请运行 'pip install openpyxl'。")
        return

    if not renaming_data:
        if output_text_signal_emit:
            output_text_signal_emit("提示：没有重命名数据可供生成Excel报告。")
        return

    excel_file_path = os.path.join(output_folder_path, report_filename)

    try:
        if not os.path.exists(output_folder_path):
            os.makedirs(output_folder_path, exist_ok=True)
            if output_text_signal_emit:
                output_text_signal_emit(f"提示：为Excel报告创建了文件夹 {output_folder_path}")
    except Exception as e:
        if output_text_signal_emit:
            output_text_signal_emit(f"错误：无法创建Excel报告文件夹 {output_folder_path}: {e}")
        return

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "商品标题"

        headers = ["Original Filename", "New Filename"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header
            sheet.column_dimensions[col_letter].width = 40

        for row_num, entry in enumerate(renaming_data, 2):
            sheet[f"A{row_num}"] = entry.get('original_name', 'N/A') # Should always have original_name
            sheet[f"B{row_num}"] = entry.get('new_name_suggestion', '') # Use empty string if suggestion is missing or empty

        excel_file_path = get_unique_filename(excel_file_path)

        workbook.save(excel_file_path)
        if output_text_signal_emit:
            output_text_signal_emit(f"成功：Excel报告已保存到 {excel_file_path}")

    except Exception as e:
        if output_text_signal_emit:
            output_text_signal_emit(f"错误：生成Excel报告失败: {e}")
